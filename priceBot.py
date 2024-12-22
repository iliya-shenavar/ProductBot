import time
import openpyxl
from telegram import Bot, InputMediaPhoto
from telegram.error import TelegramError, NetworkError
import os
import asyncio
import json
import jdatetime  

REQUIRED_LIBRARIES = [
    "openpyxl",
    "python-telegram-bot",
    "jdatetime"
]


def install_required_libraries():
    for library in REQUIRED_LIBRARIES:
        try:
            __import__(library.split("==")[0]) 
        except ImportError:
            print(f"Installing {library}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", library])
        except Exception as e:
            print(f"An error occurred while checking or installing {library}: {e}")

TELEGRAM_TOKEN = '7994752909:AAEWQ8PSpPf6whrYzXzGsckxr5R6WLHkG9g'
CHANNEL_ID = '@Pars_Ettehad'
EXCEL_FILE = 'products.xlsx'
MESSAGE_IDS_FILE = 'message_ids.json'
UPLOADED_POSTS_FILE = 'uploaded_posts.txt'

message_ids = {}
uploaded_posts = set()  


def create_excel_file(excel_file):
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Products"
        sheet.append(["کد محصول", "نام محصول", "قیمت", "آدرس عکس"])
        wb.save(excel_file)
        print(f"Excel file created: {excel_file}")
    else:
        print(f"Excel file already exists: {excel_file}")


def load_message_ids():
    global message_ids
    if os.path.exists(MESSAGE_IDS_FILE):
        with open(MESSAGE_IDS_FILE, 'r') as f:
            message_ids = json.load(f)
        print("Message IDs loaded from file.")
    else:
        print("No previous message IDs found.")


def save_message_ids():
    with open(MESSAGE_IDS_FILE, 'w') as f:
        json.dump(message_ids, f)
    print("Message IDs saved to file.")


def load_uploaded_posts():
    global uploaded_posts
    if os.path.exists(UPLOADED_POSTS_FILE):
        with open(UPLOADED_POSTS_FILE, 'r') as f:
            uploaded_posts = set(f.read().splitlines())
        print("Uploaded posts loaded from file.")
    else:
        print("No uploaded posts found.")


def save_uploaded_post(product_code):
    with open(UPLOADED_POSTS_FILE, 'a') as f:
        f.write(f"{product_code}\n")
    uploaded_posts.add(product_code)
    print(f"Uploaded post saved for product code: {product_code}")


def get_persian_date():
    return jdatetime.datetime.now().strftime('%Y/%m/%d %H:%M')


async def send_initial_price_post(product_code, product_name, price, image_url):
    if product_code in uploaded_posts:
        print(f"Skipping already uploaded product: {product_code}")
        return True  

    bot = Bot(token=TELEGRAM_TOKEN)
    persian_date = get_persian_date()
    message = f"""
کد محصول: {product_code}
نام محصول: {product_name}
قیمت مصوب: {price}
تاریخ بروزرسانی: {persian_date}
    """
    try:
        sent_message = await bot.send_photo(chat_id=CHANNEL_ID, photo=image_url, caption=message)
        message_ids[product_code] = {
            "message_id": sent_message.message_id,
            "price": price
        }
        save_message_ids()
        save_uploaded_post(product_code) 
        print(f"Initial message sent for product: {product_name}")
    except TelegramError as e:
        print(f"Error sending message for product {product_name} (Code: {product_code}): {e.message}")
        return False
    except NetworkError as e:
        print(f"Network error: {e.message}. Retrying...")
        return False
    return True


def process_price(price):
    if price is None:
        return "نامشخص"
    try:
        return f"{int(price):,} ریال"
    except (ValueError, TypeError):
        return str(price).strip()
    
async def edit_price_post(product_code, product_name, price, image_url):
    bot = Bot(token=TELEGRAM_TOKEN)
    persian_date = get_persian_date()
    message = f"""
کد محصول: {product_code}
نام محصول: {product_name}
قیمت مصوب: {price}
تاریخ بروزرسانی: {persian_date}
    """
    try:
        if product_code in message_ids:
            message_id = message_ids[product_code]["message_id"]
            await bot.edit_message_caption(chat_id=CHANNEL_ID, message_id=message_id, caption=message)
            message_ids[product_code]["price"] = price  
            save_message_ids()
            print(f"Message updated for product: {product_name}")
        else:
            print(f"No existing message found for product {product_code}. Sending new message.")
            await send_initial_price_post(product_code, product_name, price, image_url)
    except TelegramError as e:
        print(f"Error editing message for product {product_name} (Code: {product_code}): {e.message}")
    except NetworkError as e:
        print(f"Network error: {e.message}. Retrying...")


async def check_price_changes(excel_file):
    create_excel_file(excel_file)
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
        product_code = str(row[0].value).strip()
        product_name = str(row[1].value).strip()
        price = process_price(row[2].value) 
        image_url = str(row[3].value).strip()

        try:
            if product_code in message_ids:
                previous_price = message_ids[product_code].get("price", None)
                if previous_price == price:
                    print(f"Price for product {product_name} (Code: {product_code}) hasn't changed.")
                    continue
                else:
                    print(f"Price for product {product_name} (Code: {product_code}) has changed.")
                    await edit_price_post(product_code, product_name, price, image_url)
            else:
                print(f"No previous message found for product {product_name} (Code: {product_code}). Sending new message.")
                await send_initial_price_post(product_code, product_name, price, image_url)
        except Exception as e:
            print(f"Error processing product {product_name} (Code: {product_code}): {e}")
            continue  




def main():
    load_message_ids()
    load_uploaded_posts()
    create_excel_file(EXCEL_FILE)

    while True:
        try:
            asyncio.run(check_price_changes(EXCEL_FILE))
            time.sleep(5)
        except Exception as e:
            print(f"Unexpected error occurred: {e}. Retrying...")
            time.sleep(5)


if __name__ == "__main__":
    main()
